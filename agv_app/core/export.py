# -*- coding: utf-8 -*-
"""Xuất báo cáo ra Excel (.xlsx) và CSV bằng openpyxl.

Điểm nhấn theo yêu cầu "báo cáo cho sếp":
  - Sheet đầu «Cho sếp xem»: kết luận ỔN / CẦN CHÚ Ý / CẦN XỬ LÝ + vài số lớn
    + câu tiếng Việt toàn số liệu (không jargon kỹ thuật).
  - Excel vẫn GẮN CÔNG THỨC SỐNG ở các sheet chi tiết (Theo ngày…).
  - Có BIỂU ĐỒ ngay trong sheet Cho sếp xem.
  - CSV kèm cột đầu vào + chú thích công thức.

Toàn bộ chữ tiếng Việt CÓ DẤU.
"""

from __future__ import annotations

import csv
from datetime import date
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .abnormal import DEFAULT_DENOM_HOURS, DayResult, ShiftResult
from .aggregate import (
    PeriodSummary,
    aggregate_cars,
    aggregate_points,
    daynight_split,
    monthly_summaries,
    overall_summary,
    severity_buckets,
    top_api_points,
    weekday_name,
    weekday_pattern,
    weekly_summaries,
)
from .task_api_log import DayApiLogStats, csv_api_crosscheck
from .tasks import (
    DayTaskStats,
    aggregate_car_utilization,
    aggregate_models,
    shift_label_for_task,
)

# --- Style dùng chung ---------------------------------------------------------

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_BOLD = Font(bold=True, size=11)
_NORMAL = Font(size=10)
_TITLE = Font(bold=True, size=15, color="1F4E79")
_SUBTITLE = Font(size=10, italic=True, color="666666")
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_RATE_FILL = PatternFill("solid", fgColor="FCE4D6")
_KPI_FILL = PatternFill("solid", fgColor="E2EFDA")
_TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
_OK_FILL = PatternFill("solid", fgColor="C6EFCE")
_WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
_BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
_BOX_FILL = PatternFill("solid", fgColor="DDEBF7")
_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

_PCT_FMT = '0.00"%"'
_HOUR_FMT = "0.00"

SHIFT_LABEL_VI = {"day": "Ca ngày", "night": "Ca đêm"}


def _style(cell, font=_NORMAL, alignment=_CENTER, border=_THIN, fill=None, number_format=None):
    cell.font = font
    cell.alignment = alignment
    cell.border = border
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format


def _set_widths(ws: Worksheet, widths: dict):
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


def _write_header_row(ws: Worksheet, row: int, headers: List[str]) -> None:
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        _style(cell, font=_BOLD, fill=_HEADER_FILL)


# --- Sheet: Theo ngày (có công thức) ------------------------------------------

def _sheet_daily(ws: Worksheet, day_results: List[DayResult], denom_hours: float) -> int:
    """Ghi bảng theo ngày với công thức tỷ lệ. Trả về số dòng TỔNG."""
    ws.merge_cells("A1:T1")
    ws["A1"] = (
        "Chất lượng AGV theo ngày "
        "(bất thường = giờ BT / (mẫu số × số xe) × 100; "
        "稼動率 từ Tasks Log; API = taskCreate/poll từ logs)"
    )
    _style(ws["A1"], font=_TITLE, alignment=_LEFT)

    headers = [
        "Ngày", "Thứ", "Số xe", "Tổng lượt bất thường",
        "Giờ BT ca ngày", "Số xe ca ngày", "Giờ BT ca đêm", "Số xe ca đêm",
        "Số giờ mẫu số", "Tỷ lệ ca ngày (%)", "Tỷ lệ ca đêm (%)",
        "Tỷ lệ cả ngày (%)", "稼動率 (%)", "Số task", "Timeout",
        "稼動 ca ngày (%)", "稼動 ca đêm (%)",
        "API taskCreate", "API poll", "API lỗi",
    ]
    _write_header_row(ws, 3, headers)

    row = 4
    first = row
    for d in sorted(day_results, key=lambda x: x.base_date):
        ts = d.task_stats if isinstance(d.task_stats, DayTaskStats) else None
        api = d.api_log_stats if isinstance(d.api_log_stats, DayApiLogStats) else None
        ws.cell(row=row, column=1, value=d.base_date.isoformat())
        ws.cell(row=row, column=2, value=weekday_name(d.base_date))
        ws.cell(row=row, column=3, value=d.car_count)
        ws.cell(row=row, column=4, value=d.abnormal_count)
        ws.cell(row=row, column=5, value=round(d.day.abnormal_hours, 4))
        ws.cell(row=row, column=6, value=d.day.car_count)
        ws.cell(row=row, column=7, value=round(d.night.abnormal_hours, 4))
        ws.cell(row=row, column=8, value=d.night.car_count)
        ws.cell(row=row, column=9, value=denom_hours)
        ws.cell(row=row, column=10,
                value="=IF(F%d*I%d=0,0,ROUND(E%d/(I%d*F%d)*100,2))" % (row, row, row, row, row))
        ws.cell(row=row, column=11,
                value="=IF(H%d*I%d=0,0,ROUND(G%d/(I%d*H%d)*100,2))" % (row, row, row, row, row))
        ws.cell(row=row, column=12,
                value="=IF(C%d*I%d=0,0,ROUND((E%d+G%d)/(I%d*C%d)*100,2))"
                      % (row, row, row, row, row, row))
        ws.cell(row=row, column=13, value=d.utilization if d.utilization else None)
        ws.cell(row=row, column=14, value=d.task_count if d.task_count else None)
        ws.cell(row=row, column=15, value=d.timeout_count if d.timeout_count else None)
        ws.cell(row=row, column=16,
                value=(ts.utilization_day if ts and ts.day else None))
        ws.cell(row=row, column=17,
                value=(ts.utilization_night if ts and ts.night else None))
        ws.cell(row=row, column=18, value=d.api_create_count if d.api_create_count else None)
        ws.cell(row=row, column=19, value=d.api_poll_count if d.api_poll_count else None)
        ws.cell(row=row, column=20, value=d.api_error_count if d.api_error_count else None)

        for col in range(1, 21):
            fmt = _PCT_FMT if col in (10, 11, 12, 13, 16, 17) else (
                _HOUR_FMT if col in (5, 7) else None)
            fill = _RATE_FILL if col in (12, 13) else None
            _style(ws.cell(row=row, column=col), fill=fill, number_format=fmt)
        row += 1

    last = row - 1
    total = row
    summ = overall_summary(day_results, denom_hours)
    ws.cell(row=total, column=1, value="TỔNG / CHUNG")
    ws.cell(row=total, column=3, value="=SUM(C%d:C%d)" % (first, last))
    ws.cell(row=total, column=4, value="=SUM(D%d:D%d)" % (first, last))
    ws.cell(row=total, column=5, value="=SUM(E%d:E%d)" % (first, last))
    ws.cell(row=total, column=7, value="=SUM(G%d:G%d)" % (first, last))
    ws.cell(row=total, column=9, value=denom_hours)
    ws.cell(row=total, column=12,
            value="=IF(C%d*I%d=0,0,ROUND((E%d+G%d)/(I%d*C%d)*100,2))"
                  % (total, total, total, total, total, total))
    ws.cell(row=total, column=13, value=summ.utilization if summ.utilization else None)
    ws.cell(row=total, column=14, value=summ.task_count if summ.task_count else None)
    ws.cell(row=total, column=15, value=summ.timeout_count if summ.timeout_count else None)
    ws.cell(row=total, column=18, value=summ.api_create_count or None)
    ws.cell(row=total, column=19, value=summ.api_poll_count or None)
    ws.cell(row=total, column=20, value=summ.api_error_count or None)
    for col in range(1, 21):
        fmt = _PCT_FMT if col in (12, 13) else (_HOUR_FMT if col in (5, 7) else None)
        _style(ws.cell(row=total, column=col), font=_BOLD, fill=_TOTAL_FILL, number_format=fmt)

    _set_widths(ws, {
        "A": 12, "B": 10, "C": 8, "D": 16, "E": 14, "F": 13, "G": 14, "H": 13,
        "I": 13, "J": 15, "K": 15, "L": 15, "M": 12, "N": 10, "O": 10, "P": 14, "Q": 14,
        "R": 14, "S": 10, "T": 10,
    })
    ws.freeze_panes = "A4"
    return total


# --- Sheet: theo kỳ (tuần / tháng) --------------------------------------------

def _sheet_period(wb: Workbook, title: str, summaries: List[PeriodSummary],
                  denom_hours: float) -> None:
    ws = wb.create_sheet(title=title)
    ws.merge_cells("A1:M1")
    ws["A1"] = (
        title + " (bất thường = giờ BT / (mẫu số × xe-ngày) × 100; "
        "稼動率 = tổng duration / tổng (giờ ca × số xe))"
    )
    _style(ws["A1"], font=_TITLE, alignment=_LEFT)

    headers = [
        "Kỳ", "Từ ngày", "Đến ngày", "Số ngày", "Số xe (khác nhau)",
        "Số xe-ngày", "Tổng lượt bất thường", "Giờ bất thường",
        "Số giờ mẫu số", "Tỷ lệ bất thường (%)",
        "稼動率 (%)", "Số task", "Timeout",
    ]
    _write_header_row(ws, 3, headers)

    row = 4
    for s in summaries:
        ws.cell(row=row, column=1, value=s.label)
        ws.cell(row=row, column=2, value=s.date_start.isoformat() if s.date_start else "-")
        ws.cell(row=row, column=3, value=s.date_end.isoformat() if s.date_end else "-")
        ws.cell(row=row, column=4, value=s.num_days)
        ws.cell(row=row, column=5, value=s.distinct_car_count)
        ws.cell(row=row, column=6, value=s.car_days)
        ws.cell(row=row, column=7, value=s.abnormal_count)
        ws.cell(row=row, column=8, value=round(s.abnormal_hours, 4))
        ws.cell(row=row, column=9, value=denom_hours)
        ws.cell(row=row, column=10,
                value="=IF(F%d*I%d=0,0,ROUND(H%d/(I%d*F%d)*100,2))" % (row, row, row, row, row))
        util = s.utilization
        ws.cell(row=row, column=11, value=util if util else None)
        ws.cell(row=row, column=12, value=s.task_count if s.task_count else None)
        ws.cell(row=row, column=13, value=s.timeout_count if s.timeout_count else None)
        for col in range(1, 14):
            fmt = _PCT_FMT if col in (10, 11) else (_HOUR_FMT if col == 8 else None)
            fill = _RATE_FILL if col in (10, 11) else None
            _style(ws.cell(row=row, column=col), fill=fill, number_format=fmt)
        row += 1

    _set_widths(ws, {
        "A": 18, "B": 12, "C": 12, "D": 8, "E": 16, "F": 12, "G": 18,
        "H": 14, "I": 13, "J": 18, "K": 12, "L": 10, "M": 10,
    })


def _sheet_utilization(wb: Workbook, day_results: List[DayResult]) -> None:
    """Sheet 稼動率 theo ca + bảng xe (giống báo cáo tích hợp gốc)."""
    ws = wb.create_sheet(title="Tỷ lệ hoạt động")
    ws.merge_cells("A1:F1")
    ws["A1"] = "Tỷ lệ hoạt động AGV (稼動率) theo ca — từ Tasks Log CSV"
    _style(ws["A1"], font=_TITLE, alignment=_LEFT)

    row = 3
    for d in sorted(day_results, key=lambda x: x.base_date):
        ts = d.task_stats if isinstance(d.task_stats, DayTaskStats) else None
        ws.cell(row=row, column=1, value="Ngày %s (%s)" % (
            d.base_date.isoformat(), weekday_name(d.base_date)))
        _style(ws.cell(row=row, column=1), font=_BOLD, fill=_KPI_FILL, alignment=_LEFT)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

        if ts is None or not ts.has_data:
            ws.cell(row=row, column=1, value="(Không có dữ liệu Tasks Log)")
            _style(ws.cell(row=row, column=1), font=_SUBTITLE, alignment=_LEFT, border=Border())
            row += 2
            continue

        for shift_key, shift_util, label in (
            ("day", ts.day, "Ca ngày"),
            ("night", ts.night, "Ca đêm"),
        ):
            # Không mở đầu bằng "=" — openpyxl sẽ coi đó là công thức và Excel
            # phải sửa file khi mở.
            ws.cell(row=row, column=1, value="── %s ──" % label)
            _style(ws.cell(row=row, column=1), font=_BOLD, alignment=_LEFT)
            row += 1
            if shift_util is None:
                ws.cell(row=row, column=1, value="(Không có nhiệm vụ completed)")
                _style(ws.cell(row=row, column=1), font=_SUBTITLE, alignment=_LEFT, border=Border())
                row += 2
                continue

            _write_header_row(ws, row, [
                "Số xe", "Số task", "Tổng thời gian (giây)", "Tổng thời gian (giờ)",
            ])
            row += 1
            first = row
            for cr in shift_util.car_rows:
                ws.cell(row=row, column=1, value=cr.car_id)
                ws.cell(row=row, column=2, value=cr.task_count)
                ws.cell(row=row, column=3, value=cr.total_sec)
                ws.cell(row=row, column=4, value=cr.total_hours)
                for col in range(1, 5):
                    _style(ws.cell(row=row, column=col),
                           number_format=(_HOUR_FMT if col in (3, 4) else None))
                row += 1
            last = row - 1
            ws.cell(row=row, column=1, value="Tổng 稼動率")
            ws.cell(row=row, column=2, value=shift_util.utilization)
            ws.cell(row=row, column=3,
                    value="(cơ sở: %gH × %d xe)" % (shift_util.shift_h, shift_util.car_count))
            _style(ws.cell(row=row, column=1), font=_BOLD, fill=_TOTAL_FILL)
            _style(ws.cell(row=row, column=2), font=_BOLD, fill=_RATE_FILL, number_format=_PCT_FMT)
            _style(ws.cell(row=row, column=3), font=_SUBTITLE, alignment=_LEFT, fill=_TOTAL_FILL)
            row += 2

        ws.cell(row=row, column=1, value="稼動 cả ngày")
        ws.cell(row=row, column=2, value=ts.utilization_all)
        _style(ws.cell(row=row, column=1), font=_BOLD, fill=_KPI_FILL)
        _style(ws.cell(row=row, column=2), font=_BOLD, fill=_RATE_FILL, number_format=_PCT_FMT)
        row += 2

    _set_widths(ws, {"A": 18, "B": 12, "C": 22, "D": 18, "E": 12, "F": 12})


def _sheet_models(wb: Workbook, day_results: List[DayResult]) -> None:
    ws = wb.create_sheet(title="Theo model")
    ws.merge_cells("A1:K1")
    ws["A1"] = (
        "Phân bố model / sản phẩm từ Tasks Log — kèm % tổng, số ngày/xe, "
        "Min/Max chu kỳ, Timeout và % timeout"
    )
    _style(ws["A1"], font=_TITLE, alignment=_LEFT)

    task_list = [
        d.task_stats for d in day_results
        if isinstance(d.task_stats, DayTaskStats)
    ]
    models = aggregate_models(task_list)
    total_tasks = sum(m.task_count for m in models) or 1
    _write_header_row(ws, 3, [
        "Model", "Số task", "% tổng", "Số ngày", "Số xe",
        "Tổng giờ", "TB chu kỳ (giây)", "Min (phút)", "Max (phút)",
        "Timeout", "% timeout",
    ])
    row = 4
    for ms in models:
        share = round(ms.task_count / total_tasks * 100.0, 1)
        ws.cell(row=row, column=1, value=ms.model)
        ws.cell(row=row, column=2, value=ms.task_count)
        ws.cell(row=row, column=3, value=share)
        ws.cell(row=row, column=4, value=ms.day_seen_count)
        ws.cell(row=row, column=5, value=ms.car_count)
        ws.cell(row=row, column=6, value=ms.total_hours)
        ws.cell(row=row, column=7, value=ms.avg_sec)
        ws.cell(row=row, column=8, value=ms.min_min if ms.duration_count else None)
        ws.cell(row=row, column=9, value=ms.max_min if ms.duration_count else None)
        ws.cell(row=row, column=10, value=ms.timeout_count)
        ws.cell(row=row, column=11, value=ms.timeout_rate)
        for col in range(1, 12):
            _style(ws.cell(row=row, column=col),
                   number_format=(_HOUR_FMT if col in (3, 6, 7, 8, 9, 11) else None))
        row += 1
    if not models:
        ws.cell(row=4, column=1, value="(Không có dữ liệu Tasks Log)")
        _style(ws.cell(row=4, column=1), font=_SUBTITLE, alignment=_LEFT, border=Border())

    _set_widths(ws, {
        "A": 22, "B": 10, "C": 10, "D": 10, "E": 10,
        "F": 12, "G": 14, "H": 12, "I": 12, "J": 10, "K": 12,
    })

    if models:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Số task theo model (top)"
        chart.y_axis.title = "Số task"
        chart.x_axis.title = "Model"
        n = min(len(models), 15)
        data = Reference(ws, min_col=2, min_row=3, max_row=3 + n)
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + n)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.width = 18
        chart.height = 10
        ws.add_chart(chart, "G3")


def _sheet_task_cars_period(wb: Workbook, day_results: List[DayResult]) -> None:
    """稼動 theo xe gộp cả kỳ phân tích."""
    ws = wb.create_sheet(title="稼動 theo xe")
    ws.merge_cells("A1:D1")
    ws["A1"] = "Tỷ lệ hoạt động theo xe (gộp cả kỳ) — từ Tasks Log"
    _style(ws["A1"], font=_TITLE, alignment=_LEFT)

    task_list = [
        d.task_stats for d in day_results
        if isinstance(d.task_stats, DayTaskStats) and d.task_stats.has_data
    ]
    cars = aggregate_car_utilization(task_list)
    _write_header_row(ws, 3, [
        "Số xe", "Số task completed", "Tổng thời gian (giây)", "Tổng thời gian (giờ)",
    ])
    row = 4
    for cr in cars:
        ws.cell(row=row, column=1, value=cr.car_id)
        ws.cell(row=row, column=2, value=cr.task_count)
        ws.cell(row=row, column=3, value=cr.total_sec)
        ws.cell(row=row, column=4, value=cr.total_hours)
        for col in range(1, 5):
            _style(ws.cell(row=row, column=col),
                   number_format=(_HOUR_FMT if col in (3, 4) else None))
        row += 1
    if not cars:
        ws.cell(row=4, column=1, value="(Không có dữ liệu Tasks Log)")
        _style(ws.cell(row=4, column=1), font=_SUBTITLE, alignment=_LEFT, border=Border())
    else:
        total_sec = sum(c.total_sec for c in cars)
        total_tasks = sum(c.task_count for c in cars)
        ws.cell(row=row, column=1, value="TỔNG")
        ws.cell(row=row, column=2, value=total_tasks)
        ws.cell(row=row, column=3, value=round(total_sec, 1))
        ws.cell(row=row, column=4, value=round(total_sec / 3600.0, 4))
        for col in range(1, 5):
            _style(ws.cell(row=row, column=col), font=_BOLD, fill=_TOTAL_FILL,
                   number_format=(_HOUR_FMT if col in (3, 4) else None))

    _set_widths(ws, {"A": 12, "B": 18, "C": 20, "D": 18})


def _sheet_task_detail(wb: Workbook, day_results: List[DayResult],
                       settings=None) -> None:
    """Chi tiết từng nhiệm vụ Tasks Log (gộp vào cùng workbook)."""
    ws = wb.create_sheet(title="Chi tiết Tasks Log")
    ws.merge_cells("A1:J1")
    ws["A1"] = (
        "Chi tiết nhiệm vụ Tasks Log (mỗi dòng = 1 task trong cửa sổ ca ngày logic). "
        "Dùng để đối chiếu / lọc / pivot."
    )
    _style(ws["A1"], font=_TITLE, alignment=_LEFT)

    headers = [
        "Ngày logic", "Ca", "Task ID", "Số xe", "Model",
        "Gửi lúc", "Hoàn thành", "Duration (giây)", "Duration (phút)", "Trạng thái",
    ]
    _write_header_row(ws, 3, headers)

    # Ước lượng + light style nếu rất nhiều dòng
    est = sum(
        len(getattr(d.task_stats, "rows", []) or [])
        for d in day_results
        if isinstance(d.task_stats, DayTaskStats)
    )
    light = est > 5000

    row = 4
    from .abnormal import calc_shift_ranges
    from .config import load_settings
    if settings is None:
        settings = load_settings()

    for d in sorted(day_results, key=lambda x: x.base_date):
        ts = d.task_stats if isinstance(d.task_stats, DayTaskStats) else None
        if ts is None or not ts.rows:
            continue
        day_start, day_end, night_start, night_end = calc_shift_ranges(d.base_date, settings)
        for tr in ts.rows:
            shift = shift_label_for_task(
                tr.complete_time, day_start, day_end, night_start, night_end)
            shift_vi = SHIFT_LABEL_VI.get(shift, shift)
            ws.cell(row=row, column=1, value=d.base_date.isoformat())
            ws.cell(row=row, column=2, value=shift_vi)
            ws.cell(row=row, column=3, value=tr.task_id)
            ws.cell(row=row, column=4, value=tr.car_id)
            ws.cell(row=row, column=5, value=tr.model)
            ws.cell(row=row, column=6,
                    value=tr.send_time.strftime("%Y-%m-%d %H:%M:%S") if tr.send_time else "")
            ws.cell(row=row, column=7,
                    value=tr.complete_time.strftime("%Y-%m-%d %H:%M:%S") if tr.complete_time else "")
            ws.cell(row=row, column=8, value=round(tr.duration_sec, 1))
            ws.cell(row=row, column=9, value=round(tr.duration_sec / 60.0, 2))
            ws.cell(row=row, column=10, value=tr.final_state)
            if not light:
                for col in range(1, 11):
                    _style(ws.cell(row=row, column=col),
                           number_format=(_HOUR_FMT if col in (8, 9) else None))
            row += 1

    if row == 4:
        ws.cell(row=4, column=1, value="(Không có dòng Tasks Log trong kỳ)")
        _style(ws.cell(row=4, column=1), font=_SUBTITLE, alignment=_LEFT, border=Border())

    _set_widths(ws, {
        "A": 12, "B": 10, "C": 38, "D": 10, "E": 22,
        "F": 20, "G": 20, "H": 14, "I": 14, "J": 12,
    })


# --- Sheet: Điều phối API -----------------------------------------------------

def _sheet_api_dispatch(wb: Workbook, day_results: List[DayResult]) -> None:
    ws = wb.create_sheet(title="Điều phối API")
    ws.merge_cells("A1:H1")
    ws["A1"] = (
        "Điều phối MES từ Tasks API logs — số lệnh taskCreate, poll, lỗi, điểm hot"
    )
    _style(ws["A1"], font=_TITLE, alignment=_LEFT)

    _write_header_row(ws, 3, [
        "Ngày", "taskCreate", "Poll", "Task unique", "Gán xe (poll)",
        "Lỗi API (stateCode≠0)", "Điểm hot", "Số lần điểm hot",
    ])
    row = 4
    any_data = False
    for d in sorted(day_results, key=lambda x: x.base_date):
        api = d.api_log_stats if isinstance(d.api_log_stats, DayApiLogStats) else None
        if api is None or not api.has_data:
            continue
        any_data = True
        ws.cell(row=row, column=1, value=d.base_date.isoformat())
        ws.cell(row=row, column=2, value=api.create_count)
        ws.cell(row=row, column=3, value=api.poll_count)
        ws.cell(row=row, column=4, value=api.unique_tasks)
        ws.cell(row=row, column=5, value=api.assigned_car_count)
        ws.cell(row=row, column=6, value=api.api_error_count)
        ws.cell(row=row, column=7, value=api.hot_point or "")
        ws.cell(row=row, column=8, value=api.hot_point_count or None)
        for col in range(1, 9):
            _style(ws.cell(row=row, column=col))
        row += 1

    if not any_data:
        ws.cell(row=4, column=1, value="(Không có Tasks API log trong kỳ)")
        _style(ws.cell(row=4, column=1), font=_SUBTITLE, alignment=_LEFT, border=Border())
        row = 5

    # Top điểm gộp
    top_row = row + 2
    ws.cell(row=top_row, column=1, value="Top điểm trong payload API (cả kỳ)")
    _style(ws.cell(row=top_row, column=1), font=_BOLD, alignment=_LEFT, border=Border())
    _write_header_row(ws, top_row + 1, ["Điểm", "Số lần"])
    r = top_row + 2
    for pt, n in top_api_points(day_results, limit=20):
        ws.cell(row=r, column=1, value=pt)
        ws.cell(row=r, column=2, value=n)
        for col in range(1, 3):
            _style(ws.cell(row=r, column=col))
        r += 1

    _set_widths(ws, {
        "A": 14, "B": 12, "C": 10, "D": 12, "E": 14, "F": 20, "G": 12, "H": 16,
    })


def _sheet_api_create_detail(wb: Workbook, day_results: List[DayResult]) -> None:
    ws = wb.create_sheet(title="Chi tiết API taskCreate")
    ws.merge_cells("A1:G1")
    ws["A1"] = "Chi tiết từng lệnh taskCreate (point route + systemName + score)"
    _style(ws["A1"], font=_TITLE, alignment=_LEFT)

    _write_header_row(ws, 3, [
        "Ngày", "Thời điểm", "systemName", "score", "Số điểm", "Lộ trình điểm", "File nguồn",
    ])

    total_creates = 0
    for d in day_results:
        api = d.api_log_stats if isinstance(d.api_log_stats, DayApiLogStats) else None
        if api:
            total_creates += len(api.creates or [])
    light = total_creates > 5000

    row = 4
    for d in sorted(day_results, key=lambda x: x.base_date):
        api = d.api_log_stats if isinstance(d.api_log_stats, DayApiLogStats) else None
        if api is None:
            continue
        for ev in (api.creates or []):
            route = " → ".join(pt for pt, _a in ev.points) if ev.points else ""
            ws.cell(row=row, column=1, value=d.base_date.isoformat())
            ws.cell(row=row, column=2,
                    value=ev.ts.strftime("%Y-%m-%d %H:%M:%S") if ev.ts else "")
            ws.cell(row=row, column=3, value=ev.system_name)
            ws.cell(row=row, column=4, value=ev.score)
            ws.cell(row=row, column=5, value=len(ev.points))
            ws.cell(row=row, column=6, value=route)
            ws.cell(row=row, column=7, value=Path(ev.source_file).name if ev.source_file else "")
            if not light:
                for col in range(1, 8):
                    _style(ws.cell(row=row, column=col))
            row += 1

    if row == 4:
        ws.cell(row=4, column=1, value="(Không có lệnh taskCreate)")
        _style(ws.cell(row=4, column=1), font=_SUBTITLE, alignment=_LEFT, border=Border())

    _set_widths(ws, {
        "A": 12, "B": 20, "C": 12, "D": 10, "E": 10, "F": 48, "G": 28,
    })


def _sheet_csv_api_crosscheck(wb: Workbook, day_results: List[DayResult]) -> None:
    ws = wb.create_sheet(title="Đối chiếu CSV ↔ API")
    ws.merge_cells("A1:G1")
    ws["A1"] = (
        "Đối chiếu cùng ngày: số task CSV (cửa sổ ca) vs số lệnh taskCreate API"
    )
    _style(ws["A1"], font=_TITLE, alignment=_LEFT)

    _write_header_row(ws, 3, [
        "Ngày", "CSV task_count", "API create_count", "API poll",
        "Chênh (CSV − API)", "Có CSV", "Có API",
    ])
    row = 4
    rows = csv_api_crosscheck(day_results)
    if not rows:
        ws.cell(row=4, column=1, value="(Không có ngày nào có CSV hoặc API để đối chiếu)")
        _style(ws.cell(row=4, column=1), font=_SUBTITLE, alignment=_LEFT, border=Border())
    else:
        for r in rows:
            ws.cell(row=row, column=1, value=r["date"].isoformat())  # type: ignore[union-attr]
            ws.cell(row=row, column=2, value=r["csv_task_count"])
            ws.cell(row=row, column=3, value=r["api_create_count"])
            ws.cell(row=row, column=4, value=r["api_poll_count"])
            ws.cell(row=row, column=5, value=r["diff_csv_minus_api"])
            ws.cell(row=row, column=6, value="Có" if r["has_csv"] else "Không")
            ws.cell(row=row, column=7, value="Có" if r["has_api"] else "Không")
            for col in range(1, 8):
                _style(ws.cell(row=row, column=col))
            row += 1

    _set_widths(ws, {
        "A": 12, "B": 14, "C": 16, "D": 10, "E": 16, "F": 10, "G": 10,
    })


# --- Sheet: theo điểm ---------------------------------------------------------

def _sheet_points(wb: Workbook, day_results: List[DayResult]) -> Worksheet:
    ws = wb.create_sheet(title="Theo điểm")
    ws.merge_cells("A1:I1")
    ws["A1"] = (
        "Xếp hạng ĐIỂM hay kẹt — kèm phân ca, số ngày xuất hiện, "
        "TB/Max để phân biệt kẹt nhiều lần ngắn vs ít lần nhưng nặng"
    )
    _style(ws["A1"], font=_TITLE, alignment=_LEFT)

    _write_header_row(ws, 3, [
        "Điểm", "Số lượt", "Số xe", "Số ngày",
        "Ca ngày", "Ca đêm",
        "Tổng thời gian (phút)", "Trung bình (phút)", "Max (phút)",
    ])
    row = 4
    for ps in aggregate_points(day_results):
        ws.cell(row=row, column=1, value="Điểm %s" % ps.point_id)
        ws.cell(row=row, column=2, value=ps.abnormal_count)
        ws.cell(row=row, column=3, value=ps.car_count)
        ws.cell(row=row, column=4, value=ps.day_seen_count)
        ws.cell(row=row, column=5, value=ps.day_count)
        ws.cell(row=row, column=6, value=ps.night_count)
        ws.cell(row=row, column=7, value=ps.abnormal_min)
        ws.cell(row=row, column=8,
                value="=IF(B%d=0,0,ROUND(G%d/B%d,2))" % (row, row, row))
        ws.cell(row=row, column=9, value=ps.max_min)
        for col in range(1, 10):
            _style(ws.cell(row=row, column=col),
                   number_format=(_HOUR_FMT if col in (7, 8, 9) else None))
        row += 1

    _set_widths(ws, {
        "A": 14, "B": 12, "C": 10, "D": 10, "E": 10, "F": 10,
        "G": 20, "H": 16, "I": 12,
    })
    return ws


# --- Sheet: theo xe -----------------------------------------------------------

def _sheet_cars(wb: Workbook, day_results: List[DayResult]) -> Worksheet:
    ws = wb.create_sheet(title="Theo xe")
    ws.merge_cells("A1:J1")
    ws["A1"] = (
        "Xếp hạng XE bất thường — kèm số ngày, số điểm, điểm hay kẹt, "
        "phân ca, TB/Max để phân biệt xe hay kẹt ngắn vs dừng nặng"
    )
    _style(ws["A1"], font=_TITLE, alignment=_LEFT)

    _write_header_row(ws, 3, [
        "Số xe", "Số lượt", "Số ngày", "Số điểm", "Điểm hay kẹt",
        "Ca ngày", "Ca đêm",
        "Tổng thời gian (phút)", "Trung bình (phút)", "Max (phút)",
    ])
    row = 4
    for cs in aggregate_cars(day_results):
        hot = ("Điểm %s (%d)" % (cs.hot_point, cs.hot_point_count)
               if cs.hot_point else "-")
        ws.cell(row=row, column=1, value=cs.car_id)
        ws.cell(row=row, column=2, value=cs.abnormal_count)
        ws.cell(row=row, column=3, value=cs.day_seen_count)
        ws.cell(row=row, column=4, value=cs.point_count)
        ws.cell(row=row, column=5, value=hot)
        ws.cell(row=row, column=6, value=cs.day_count)
        ws.cell(row=row, column=7, value=cs.night_count)
        ws.cell(row=row, column=8, value=cs.abnormal_min)
        ws.cell(row=row, column=9,
                value="=IF(B%d=0,0,ROUND(H%d/B%d,2))" % (row, row, row))
        ws.cell(row=row, column=10, value=cs.max_min)
        for col in range(1, 11):
            _style(ws.cell(row=row, column=col),
                   number_format=(_HOUR_FMT if col in (8, 9, 10) else None))
        row += 1

    _set_widths(ws, {
        "A": 12, "B": 10, "C": 10, "D": 10, "E": 18,
        "F": 10, "G": 10, "H": 20, "I": 16, "J": 12,
    })
    return ws


# --- Sheet: chi tiết bất thường -----------------------------------------------

def _sheet_detail(wb: Workbook, day_results: List[DayResult]) -> None:
    ws = wb.create_sheet(title="Chi tiết bất thường")
    ws.merge_cells("A1:G1")
    ws["A1"] = "Chi tiết từng lần dừng bất thường"
    _style(ws["A1"], font=_TITLE, alignment=_LEFT)

    _write_header_row(ws, 3, [
        "Ngày", "Thứ", "Ca", "Số xe", "Điểm", "Giờ đến", "Thời gian dừng (phút)",
    ])
    # Ước lượng số dòng; >5k thì bỏ style từng ô để xuất cả tháng nhanh hơn
    est = sum(d.abnormal_count for d in day_results)
    light = est > 5000

    row = 4
    first = row
    for d in sorted(day_results, key=lambda x: x.base_date):
        wd = weekday_name(d.base_date)
        for shift in (d.day, d.night):
            by_car = shift.abnormal_by_car()
            shift_label = SHIFT_LABEL_VI.get(shift.label, shift.label)
            for car_id in sorted(by_car.keys()):
                for rec in by_car[car_id]:
                    ws.cell(row=row, column=1, value=d.base_date.isoformat())
                    ws.cell(row=row, column=2, value=wd)
                    ws.cell(row=row, column=3, value=shift_label)
                    ws.cell(row=row, column=4, value=car_id)
                    ws.cell(row=row, column=5, value="Điểm %s" % rec.point_id)
                    ws.cell(row=row, column=6, value=rec.arrival_time.strftime("%H:%M:%S"))
                    c7 = ws.cell(row=row, column=7, value=rec.stay_min)
                    if light:
                        c7.number_format = _HOUR_FMT
                    else:
                        for col in range(1, 8):
                            _style(ws.cell(row=row, column=col),
                                   number_format=(_HOUR_FMT if col == 7 else None))
                    row += 1

    last = row - 1
    if last >= first:
        ws.cell(row=row, column=6, value="Tổng thời gian (phút)")
        ws.cell(row=row, column=7, value="=SUM(G%d:G%d)" % (first, last))
        _style(ws.cell(row=row, column=6), font=_BOLD, fill=_TOTAL_FILL, alignment=_LEFT)
        _style(ws.cell(row=row, column=7), font=_BOLD, fill=_TOTAL_FILL, number_format=_HOUR_FMT)
    else:
        ws.cell(row=row, column=1, value="(Không có lần bất thường nào)")
        _style(ws.cell(row=row, column=1), font=_SUBTITLE, alignment=_LEFT, border=Border())

    _set_widths(ws, {"A": 12, "B": 10, "C": 10, "D": 8, "E": 12, "F": 12, "G": 20})
    ws.freeze_panes = "A4"


# --- Sheet: Tổng quan (dành cho lãnh đạo — nhìn số là hiểu) --------------------

def _rate_status(rate: float, rate_ok: float, rate_warn: float):
    """Trả về (nhãn, màu chữ, fill) theo ngưỡng tỷ lệ kẹt."""
    if rate >= rate_warn:
        return ("CẦN XỬ LÝ", "9C0006", _BAD_FILL)
    if rate >= rate_ok:
        return ("CẦN CHÚ Ý", "9C5700", _WARN_FILL)
    return ("ỔN", "006100", _OK_FILL)


def _boss_findings(day_results: List[DayResult], summ, denom_hours: float,
                   threshold_min: int, rate_ok: float, rate_warn: float) -> List[str]:
    """Vài câu tiếng Việt ngắn, toàn số — sếp đọc là biết đang sao."""
    lines: List[str] = []
    rate = summ.abnormal_rate
    status, _, _ = _rate_status(rate, rate_ok, rate_warn)

    lines.append(
        "1) Kết luận: %s — tỷ lệ kẹt xe cả kỳ = %.2f%% "
        "(dưới %.0f%% = ổn, %.0f–%.0f%% = chú ý, trên %.0f%% = cần xử lý)."
        % (status, rate, rate_ok, rate_ok, rate_warn, rate_warn)
    )

    days_agv = summ.days_with_agv
    if days_agv < summ.num_days:
        lines.append(
            "2) Dữ liệu: có %d/%d ngày thiếu Log AGV nên số kẹt chỉ tính trên "
            "%d ngày có log (ngày chỉ có Tasks CSV không đếm là «0 kẹt»)."
            % (summ.num_days - days_agv, summ.num_days, days_agv)
        )
    else:
        lines.append(
            "2) Dữ liệu: đủ Log AGV cho cả %d ngày trong báo cáo."
            % summ.num_days
        )

    if summ.top_point and summ.top_point.abnormal_count > 0:
        ps = summ.top_point
        lines.append(
            "3) Điểm kẹt nhiều nhất: Điểm %s — %d lần, tổng %.0f phút, "
            "TB %.1f phút/lần, Max %.0f phút (%d xe)."
            % (ps.point_id, ps.abnormal_count, ps.abnormal_min,
               ps.avg_min, ps.max_min, ps.car_count)
        )
    else:
        lines.append("3) Điểm kẹt: không ghi nhận lượt dừng vượt ngưỡng.")

    if summ.top_car and summ.top_car.abnormal_count > 0:
        cs = summ.top_car
        hot = ("hay kẹt ở Điểm %s" % cs.hot_point) if cs.hot_point else "nhiều điểm"
        lines.append(
            "4) Xe cần xem: Xe %s — %d lần kẹt, tổng %.0f phút "
            "(TB %.1f phút/lần, Max %.0f phút) — %s."
            % (cs.car_id, cs.abnormal_count, cs.abnormal_min,
               cs.avg_min, cs.max_min, hot)
        )
    else:
        lines.append("4) Xe: chưa có xe nào vượt ngưỡng dừng.")

    if summ.worst_day is not None:
        wd = summ.worst_day
        lines.append(
            "5) Ngày tệ nhất: %s (%s) — tỷ lệ kẹt %.2f%%, %d lần dừng."
            % (wd.base_date.isoformat(), weekday_name(wd.base_date),
               wd.abnormal_rate(denom_hours), wd.abnormal_count)
        )

    if summ.utilization > 0:
        util_note = "cao" if summ.utilization >= 60 else (
            "trung bình" if summ.utilization >= 40 else "thấp")
        lines.append(
            "6) Tỷ lệ chạy việc (稼動): %.2f%% (%s) — %d nhiệm vụ, %d timeout."
            % (summ.utilization, util_note, summ.task_count, summ.timeout_count)
        )
    else:
        lines.append(
            "6) Tỷ lệ chạy việc: chưa có Tasks Log — chỉ xem được phần kẹt xe.")

    lines.append(
        "7) Quy ước: «kẹt» = xe dừng tại điểm quá %d phút (không tính sạc)."
        % threshold_min
    )
    return lines


def _sheet_dashboard(ws: Worksheet, day_results: List[DayResult], denom_hours: float,
                     daily_total_row: int, point_last_row: int,
                     settings=None) -> None:
    """Sheet đầu — sếp chỉ cần mở sheet này."""
    from .config import DEFAULT_RATE_OK, DEFAULT_RATE_WARN, DEFAULT_THRESHOLD_MIN

    rate_ok = float(getattr(settings, "rate_ok", DEFAULT_RATE_OK) if settings else DEFAULT_RATE_OK)
    rate_warn = float(getattr(settings, "rate_warn", DEFAULT_RATE_WARN) if settings else DEFAULT_RATE_WARN)
    threshold_min = int(
        getattr(settings, "threshold_min", DEFAULT_THRESHOLD_MIN) if settings else DEFAULT_THRESHOLD_MIN)

    summ = overall_summary(day_results, denom_hours)
    rate = summ.abnormal_rate
    status, status_color, status_fill = _rate_status(rate, rate_ok, rate_warn)

    # --- Tiêu đề ---
    ws.merge_cells("A1:G1")
    ws["A1"] = "BÁO CÁO NHANH CHO SẾP — CHẤT LƯỢNG AGV"
    _style(ws["A1"], font=Font(bold=True, size=18, color="1F4E79"),
           alignment=_LEFT, border=Border())

    ws.merge_cells("A2:G2")
    date_range = "-"
    if day_results:
        d0 = min(d.base_date for d in day_results)
        d1 = max(d.base_date for d in day_results)
        date_range = "%s → %s" % (d0.isoformat(), d1.isoformat())
    ws["A2"] = (
        "Kỳ: %s   ·   %d ngày   ·   Chỉ cần đọc sheet này. "
        "Các sheet sau là chi tiết cho kỹ thuật."
        % (date_range, summ.num_days)
    )
    _style(ws["A2"], font=_SUBTITLE, alignment=_LEFT, border=Border())

    # --- Khối kết luận to ---
    ws.merge_cells("A4:B4")
    ws["A4"] = "KẾT LUẬN"
    _style(ws["A4"], font=Font(bold=True, size=12, color="FFFFFF"),
           fill=PatternFill("solid", fgColor="1F4E79"), alignment=_CENTER)

    ws.merge_cells("C4:E4")
    ws["C4"] = status
    _style(ws["C4"], font=Font(bold=True, size=16, color=status_color),
           fill=status_fill, alignment=_CENTER)

    ws.merge_cells("F4:G4")
    ws["F4"] = "Tỷ lệ kẹt xe: %.2f%%" % rate
    _style(ws["F4"], font=Font(bold=True, size=14, color=status_color),
           fill=status_fill, alignment=_CENTER)

    # --- 6 ô số lớn (ngôn ngữ đời thường) ---
    big_kpis = [
        ("Tỷ lệ kẹt xe (%)", rate, "Thấp hơn càng tốt", status_fill),
        ("Tổng lần kẹt", summ.abnormal_count, "Số lần dừng quá %d phút" % threshold_min, _BOX_FILL),
        ("Tổng phút kẹt", round(summ.abnormal_hours * 60.0, 0), "Cộng dồn cả kỳ", _BOX_FILL),
        ("Tỷ lệ chạy việc (%)", summ.utilization if summ.utilization else "-",
         "稼動 — cao hơn càng tốt", _KPI_FILL),
        ("Số xe", summ.distinct_cars, "Xe khác nhau có dữ liệu", _BOX_FILL),
        ("Số nhiệm vụ", summ.task_count if summ.task_count else "-",
         "Timeout: %d" % summ.timeout_count, _KPI_FILL),
    ]
    r = 6
    for i, (label, value, hint, fill) in enumerate(big_kpis):
        col = 1 + (i % 3) * 2
        if i > 0 and i % 3 == 0:
            r += 3
        lc = ws.cell(row=r, column=col, value=label)
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
        _style(lc, font=Font(bold=True, size=10, color="1F4E79"), fill=fill, alignment=_CENTER)

        vc = ws.cell(row=r + 1, column=col, value=value)
        ws.merge_cells(start_row=r + 1, start_column=col, end_row=r + 1, end_column=col + 1)
        _style(vc, font=Font(bold=True, size=18, color="C00000"), fill=fill, alignment=_CENTER,
               number_format=_PCT_FMT if isinstance(value, float) and " (%)" in label else None)

        hc = ws.cell(row=r + 2, column=col, value=hint)
        ws.merge_cells(start_row=r + 2, start_column=col, end_row=r + 2, end_column=col + 1)
        _style(hc, font=Font(size=9, italic=True, color="666666"), fill=fill, alignment=_CENTER)

    # --- Sếp cần biết gì ---
    r += 4
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="SẾP CẦN BIẾT GÌ (đọc từng dòng — toàn là số)")
    _style(ws.cell(row=r, column=1),
           font=Font(bold=True, size=12, color="FFFFFF"),
           fill=PatternFill("solid", fgColor="1F4E79"), alignment=_LEFT)

    for line in _boss_findings(day_results, summ, denom_hours, threshold_min, rate_ok, rate_warn):
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        cell = ws.cell(row=r, column=1, value=line)
        _style(cell, font=Font(size=11), alignment=_LEFT, fill=_BOX_FILL)

    # --- Bảng Top cần xử lý ---
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="TOP CẦN XỬ LÝ (ưu tiên từ trên xuống)")
    _style(ws.cell(row=r, column=1),
           font=Font(bold=True, size=12, color="FFFFFF"),
           fill=PatternFill("solid", fgColor="C00000"), alignment=_LEFT)

    r += 1
    _write_header_row(ws, r, [
        "Hạng", "Loại", "Đối tượng", "Số lần", "Tổng phút", "TB phút/lần", "Ghi chú nhanh",
    ])

    points = aggregate_points(day_results)[:5]
    cars = aggregate_cars(day_results)[:5]
    top_rows = []
    for i, ps in enumerate(points, 1):
        top_rows.append((
            i, "Điểm hay kẹt", "Điểm %s" % ps.point_id,
            ps.abnormal_count, ps.abnormal_min, ps.avg_min,
            "Max %.0f phút · %d xe · %d ngày" % (ps.max_min, ps.car_count, ps.day_seen_count),
        ))
    for i, cs in enumerate(cars, 1):
        top_rows.append((
            i, "Xe hay kẹt", "Xe %s" % cs.car_id,
            cs.abnormal_count, cs.abnormal_min, cs.avg_min,
            "Max %.0f phút · hay kẹt Điểm %s" % (
                cs.max_min, cs.hot_point or "-"),
        ))
    # Ưu tiên: tổng phút giảm dần để sếp thấy chỗ tốn thời gian nhất
    top_rows.sort(key=lambda x: (-x[4], -x[3]))
    top_rows = top_rows[:8]

    for row_data in top_rows:
        r += 1
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            accent = _BAD_FILL if (c == 5 and isinstance(val, (int, float)) and val >= 60) else (
                _WARN_FILL if (c == 5 and isinstance(val, (int, float)) and val >= 30) else None)
            _style(cell, number_format=(_HOUR_FMT if c in (5, 6) else None), fill=accent)

    if not top_rows:
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws.cell(row=r, column=1, value="Không có điểm/xe vượt ngưỡng trong kỳ này.")
        _style(ws.cell(row=r, column=1), font=_SUBTITLE, alignment=_LEFT, fill=_OK_FILL)

    # --- Ngày cần xem ---
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="NGÀY CẦN XEM (tỷ lệ kẹt cao nhất)")
    _style(ws.cell(row=r, column=1),
           font=Font(bold=True, size=12, color="FFFFFF"),
           fill=PatternFill("solid", fgColor="1F4E79"), alignment=_LEFT)

    r += 1
    _write_header_row(ws, r, [
        "Ngày", "Thứ", "Số lần kẹt", "Tỷ lệ kẹt (%)", "Tỷ lệ chạy việc (%)",
        "Số nhiệm vụ", "Mức",
    ])
    ranked_days = sorted(
        [d for d in day_results
         if getattr(d, "has_agv_log", True) and d.car_count > 0],
        key=lambda d: (-d.abnormal_rate(denom_hours), -d.abnormal_count),
    )[:7]
    for d in ranked_days:
        r += 1
        d_rate = d.abnormal_rate(denom_hours)
        st, _, fill = _rate_status(d_rate, rate_ok, rate_warn)
        vals = [
            d.base_date.isoformat(),
            weekday_name(d.base_date),
            d.abnormal_count,
            d_rate,
            d.utilization if d.utilization else "-",
            d.task_count if d.task_count else "-",
            st,
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            _style(cell, fill=fill if c in (4, 7) else None,
                   number_format=(_PCT_FMT if c == 4 else None))

    if not ranked_days:
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws.cell(row=r, column=1,
                value="Không có ngày nào có Log AGV để xếp hạng tỷ lệ kẹt.")
        _style(ws.cell(row=r, column=1), font=_SUBTITLE, alignment=_LEFT, fill=_WARN_FILL)

    _set_widths(ws, {
        "A": 14, "B": 14, "C": 16, "D": 16, "E": 14, "F": 14, "G": 36,
    })
    ws.row_dimensions[4].height = 28

    # --- Biểu đồ (đặt dưới, tiêu đề đời thường) ---
    chart_row = r + 2
    ws_daily = ws.parent["Theo ngày"]
    daily_last_data = daily_total_row - 1
    if daily_last_data >= 4:
        chart = LineChart()
        chart.title = "Tỷ lệ kẹt xe theo ngày (%) — càng thấp càng tốt"
        chart.height = 8
        chart.width = 18
        chart.y_axis.title = "%"
        data = Reference(ws_daily, min_col=12, min_row=3, max_row=daily_last_data)
        cats = Reference(ws_daily, min_col=1, min_row=4, max_row=daily_last_data)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "A%d" % chart_row)

    if point_last_row >= 4:
        ws_points = ws.parent["Theo điểm"]
        top_last = min(point_last_row, 13)
        bar = BarChart()
        bar.type = "bar"
        bar.title = "Top điểm hay kẹt (số lần)"
        bar.height = 8
        bar.width = 18
        data = Reference(ws_points, min_col=2, min_row=3, max_row=top_last)
        cats = Reference(ws_points, min_col=1, min_row=4, max_row=top_last)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        ws.add_chart(bar, "A%d" % (chart_row + 16))


# --- Điểm vào chính -----------------------------------------------------------

def export_full_report(day_results: List[DayResult], out_path: Path,
                       threshold_min: int = 12,
                       denom_hours: float = DEFAULT_DENOM_HOURS,
                       settings=None) -> Path:
    """Xuất workbook tổng hợp nhiều sheet (công thức sống + biểu đồ + Tasks Log)."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Cho sếp xem"

    ws_daily = wb.create_sheet(title="Theo ngày")
    daily_total_row = _sheet_daily(ws_daily, day_results, denom_hours)

    _sheet_period(wb, "Theo tuần", weekly_summaries(day_results, denom_hours), denom_hours)
    _sheet_period(wb, "Theo tháng", monthly_summaries(day_results, denom_hours), denom_hours)

    _sheet_utilization(wb, day_results)
    _sheet_models(wb, day_results)
    _sheet_task_cars_period(wb, day_results)
    _sheet_task_detail(wb, day_results, settings=settings)

    _sheet_api_dispatch(wb, day_results)
    _sheet_api_create_detail(wb, day_results)
    _sheet_csv_api_crosscheck(wb, day_results)

    ws_points = _sheet_points(wb, day_results)
    point_last_row = ws_points.max_row
    _sheet_cars(wb, day_results)
    _sheet_detail(wb, day_results)

    _sheet_dashboard(
        ws_dash, day_results, denom_hours, daily_total_row, point_last_row,
        settings=settings)

    wb.save(out_path)
    return out_path


# --- Định dạng gốc (parity với ptich_agv, tỷ lệ là công thức) -----------------

def write_original_shift_sheet(ws: Worksheet, shift: ShiftResult, shift_label_vi: str,
                               threshold_min: int, denom_hours: float = DEFAULT_DENOM_HOURS) -> None:
    """Ghi một ca theo đúng bố cục xlsx bản gốc: B=xe, C=số lần, D=điểm/giờ, E=phút.

    Ô tổng tỷ lệ (C3) dùng CÔNG THỨC thay vì chuỗi tĩnh để đối chiếu được.
    """
    ws.merge_cells("B1:E1")
    c = ws["B1"]
    c.value = "AGV xe - Thống kê thời gian bất thường (%s)" % shift_label_vi
    _style(c, font=Font(bold=True, size=14))

    ws["B3"] = "Tổng tỷ lệ bất thường (O)"
    _style(ws["B3"], font=_BOLD)
    ws.merge_cells("D3:E3")
    ws["D3"] = "*Tổng (E) của mọi xe / (số giờ mẫu số × tổng số xe)"
    _style(ws["D3"], font=Font(size=9, italic=True))

    ws["C5"] = "Số lần dừng bất thường"
    ws["D5"] = "Điểm bất thường"
    ws["E5"] = "Thời gian lưu bất thường"
    for col in ("C", "D", "E"):
        _style(ws["%s5" % col], font=_BOLD, fill=_HEADER_FILL)

    ws["B7"] = "Số xe \\ Định nghĩa"
    ws["C7"] = "Dừng quá %d phút (trừ sạc)" % threshold_min
    ws["D7"] = "Ghi chú điểm bất thường (1 điểm/1 giờ/1 dòng)"
    ws["E7"] = "Thời gian lưu tại điểm bất thường (phút)"
    for col in ("B", "C", "D", "E"):
        _style(ws["%s7" % col], font=_BOLD, fill=_HEADER_FILL)

    by_car = shift.abnormal_by_car()
    all_cars = shift.all_cars
    row = 8
    first = row
    for car_id in all_cars:
        records = by_car.get(car_id, [])
        if records:
            for j, rec in enumerate(records):
                if j == 0:
                    ws.cell(row=row, column=2, value=car_id)
                    ws.cell(row=row, column=3, value=len(records))
                ws.cell(row=row, column=4,
                        value="Điểm%s / %s" % (rec.point_id, rec.arrival_time.strftime("%H:%M:%S")))
                ws.cell(row=row, column=5, value=rec.stay_min)
                for col in range(2, 6):
                    _style(ws.cell(row=row, column=col),
                           number_format=(_HOUR_FMT if col == 5 else None))
                row += 1
        else:
            ws.cell(row=row, column=2, value=car_id)
            ws.cell(row=row, column=3, value=0)
            ws.cell(row=row, column=4, value="-")
            ws.cell(row=row, column=5, value="-")
            for col in range(2, 6):
                _style(ws.cell(row=row, column=col))
            row += 1
    last = row - 1

    car_count = len(all_cars)
    if car_count > 0 and last >= first:
        ws["C3"] = ("=ROUND(SUM(E%d:E%d)/60/(%g*%d)*100,2)"
                    % (first, last, denom_hours, car_count))
    else:
        ws["C3"] = 0
    _style(ws["C3"], font=Font(bold=True, size=11, color="FF0000"), fill=_RATE_FILL,
           number_format=_PCT_FMT)

    note_row = row + 2
    ws.cell(row=note_row, column=2,
            value="Ngưỡng bất thường: %d phút (trừ sạc)" % threshold_min)
    _style(ws.cell(row=note_row, column=2),
           font=Font(size=9, italic=True, color="666666"),
           alignment=_LEFT, border=Border())

    _set_widths(ws, {"A": 2, "B": 12, "C": 20, "D": 30, "E": 22})


def export_original_shift(shift: ShiftResult, base_date: date, out_dir: Path,
                          threshold_min: int, shift_suffix: str,
                          denom_hours: float = DEFAULT_DENOM_HOURS) -> Path:
    """Tái tạo file Log{date}_{day|night}.xlsx theo định dạng gốc."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / ("Log%s_%s.xlsx" % (base_date.strftime("%Y%m%d"), shift_suffix))
    wb = Workbook()
    ws = wb.active
    ws.title = "daily"
    write_original_shift_sheet(ws, shift, SHIFT_LABEL_VI.get(shift.label, shift.label),
                               threshold_min, denom_hours)
    wb.save(out_file)
    return out_file


# --- CSV (giá trị + đầu vào + chú thích công thức) ----------------------------

def export_summary_csv(day_results: List[DayResult], out_path: Path,
                       denom_hours: float = DEFAULT_DENOM_HOURS,
                       settings=None) -> Path:
    """Xuất CSV tổng quan.

    Vì CSV không giữ được công thức sống ổn định (Excel tiếng Việt dùng ';'), file này
    kèm ĐẦY ĐỦ CỘT ĐẦU VÀO để tự tính lại + khối chú thích công thức bằng chữ.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)

        w.writerow(["BÁO CÁO CHẤT LƯỢNG HOẠT ĐỘNG AGV (tổng quan theo ngày)"])
        w.writerow([])

        # Khối chú thích công thức
        w.writerow(["CÔNG THỨC TÍNH"])
        w.writerow(["Tỷ lệ ca ngày (%)", "= Giờ BT ca ngày / (Số giờ mẫu số × Số xe ca ngày) × 100"])
        w.writerow(["Tỷ lệ ca đêm (%)", "= Giờ BT ca đêm / (Số giờ mẫu số × Số xe ca đêm) × 100"])
        w.writerow(["Tỷ lệ cả ngày (%)",
                    "= (Giờ BT ca ngày + Giờ BT ca đêm) / (Số giờ mẫu số × Số xe) × 100"])
        w.writerow(["稼動率 ca (%)",
                    "= Tổng duration completed / (thời lượng ca × số xe có task) × 100"])
        w.writerow(["Số giờ mẫu số hiện dùng", denom_hours])
        w.writerow([])

        # Bảng dữ liệu (kèm cột đầu vào để kiểm chứng)
        w.writerow([
            "Ngày", "Thứ", "Số xe", "Tổng lượt bất thường",
            "Giờ BT ca ngày", "Số xe ca ngày", "Giờ BT ca đêm", "Số xe ca đêm",
            "Số giờ mẫu số",
            "Tỷ lệ ca ngày (%)", "Tỷ lệ ca đêm (%)", "Tỷ lệ cả ngày (%)",
            "稼動率 (%)", "稼動 ca ngày (%)", "稼動 ca đêm (%)",
            "Số task", "Timeout", "Số file log",
            "API taskCreate", "API poll", "API lỗi",
        ])
        for d in sorted(day_results, key=lambda x: x.base_date):
            ts = d.task_stats if isinstance(d.task_stats, DayTaskStats) else None
            w.writerow([
                d.base_date.isoformat(),
                weekday_name(d.base_date),
                d.car_count,
                d.abnormal_count,
                round(d.day.abnormal_hours, 4),
                d.day.car_count,
                round(d.night.abnormal_hours, 4),
                d.night.car_count,
                denom_hours,
                d.day.abnormal_rate(denom_hours),
                d.night.abnormal_rate(denom_hours),
                d.abnormal_rate(denom_hours),
                d.utilization,
                ts.utilization_day if ts else "",
                ts.utilization_night if ts else "",
                d.task_count,
                d.timeout_count,
                d.log_file_count,
                d.api_create_count,
                d.api_poll_count,
                d.api_error_count,
            ])

            # Khối 稼動 theo ca (giống reportAnalyse)
            if ts and ts.has_data:
                for label, su in (("Ca ngày", ts.day), ("Ca đêm", ts.night)):
                    w.writerow([])
                    w.writerow(["【稼動率】 %s %s" % (d.base_date.isoformat(), label)])
                    if su is None:
                        w.writerow(["(Không có dữ liệu)"])
                        continue
                    w.writerow(["Số xe", "Số task", "Tổng giây", "Tổng giờ"])
                    for cr in su.car_rows:
                        w.writerow([cr.car_id, cr.task_count, cr.total_sec, cr.total_hours])
                    w.writerow(["Tổng 稼動率", "%s%%" % su.utilization,
                                "(cơ sở: %gH × %d xe)" % (su.shift_h, su.car_count)])

        # Model
        w.writerow([])
        w.writerow(["THEO MODEL (Tasks Log)"])
        w.writerow(["Model", "Số task", "% tổng", "Số ngày", "Số xe",
                    "Tổng giờ", "TB chu kỳ (giây)", "Min (phút)", "Max (phút)",
                    "Timeout", "% timeout"])
        task_list = [
            d.task_stats for d in day_results
            if isinstance(d.task_stats, DayTaskStats)
        ]
        models = aggregate_models(task_list)
        total_tasks = sum(m.task_count for m in models) or 1
        for ms in models:
            share = round(ms.task_count / total_tasks * 100.0, 1)
            w.writerow([
                ms.model, ms.task_count, share, ms.day_seen_count, ms.car_count,
                ms.total_hours, ms.avg_sec,
                ms.min_min if ms.duration_count else "",
                ms.max_min if ms.duration_count else "",
                ms.timeout_count, ms.timeout_rate,
            ])

        # 稼動 theo xe (kỳ)
        w.writerow([])
        w.writerow(["稼動 THEO XE (GỘP CẢ KỲ)"])
        w.writerow(["Số xe", "Số task", "Tổng giây", "Tổng giờ"])
        for cr in aggregate_car_utilization(task_list):
            w.writerow([cr.car_id, cr.task_count, cr.total_sec, cr.total_hours])

        # Chi tiết từng task
        w.writerow([])
        w.writerow(["CHI TIẾT TASKS LOG"])
        w.writerow([
            "Ngày logic", "Ca", "Task ID", "Số xe", "Model",
            "Gửi lúc", "Hoàn thành", "Duration (giây)", "Trạng thái",
        ])
        from .abnormal import calc_shift_ranges
        from .config import load_settings
        if settings is None:
            settings = load_settings()
        for d in sorted(day_results, key=lambda x: x.base_date):
            ts = d.task_stats if isinstance(d.task_stats, DayTaskStats) else None
            if ts is None or not ts.rows:
                continue
            day_start, day_end, night_start, night_end = calc_shift_ranges(
                d.base_date, settings)
            for tr in ts.rows:
                shift = shift_label_for_task(
                    tr.complete_time, day_start, day_end, night_start, night_end)
                w.writerow([
                    d.base_date.isoformat(),
                    SHIFT_LABEL_VI.get(shift, shift),
                    tr.task_id,
                    tr.car_id,
                    tr.model,
                    tr.send_time.strftime("%Y-%m-%d %H:%M:%S") if tr.send_time else "",
                    tr.complete_time.strftime("%Y-%m-%d %H:%M:%S") if tr.complete_time else "",
                    round(tr.duration_sec, 1),
                    tr.final_state,
                ])

        # Xếp hạng điểm
        w.writerow([])
        w.writerow(["XẾP HẠNG ĐIỂM HAY KẸT"])
        w.writerow(["Điểm", "Số lượt", "Số xe", "Số ngày", "Ca ngày", "Ca đêm",
                    "Tổng thời gian (phút)", "Trung bình (phút)", "Max (phút)"])
        for ps in aggregate_points(day_results):
            w.writerow(["Điểm %s" % ps.point_id, ps.abnormal_count, ps.car_count,
                        ps.day_seen_count, ps.day_count, ps.night_count,
                        ps.abnormal_min, ps.avg_min, ps.max_min])

        # Xếp hạng xe
        w.writerow([])
        w.writerow(["XẾP HẠNG XE BẤT THƯỜNG"])
        w.writerow(["Số xe", "Số lượt", "Số ngày", "Số điểm", "Điểm hay kẹt",
                    "Ca ngày", "Ca đêm",
                    "Tổng thời gian (phút)", "Trung bình (phút)", "Max (phút)"])
        for cs in aggregate_cars(day_results):
            hot = ("Điểm %s (%d)" % (cs.hot_point, cs.hot_point_count)
                   if cs.hot_point else "-")
            w.writerow([cs.car_id, cs.abnormal_count, cs.day_seen_count, cs.point_count,
                        hot, cs.day_count, cs.night_count,
                        cs.abnormal_min, cs.avg_min, cs.max_min])

        # So sánh ca ngày / đêm
        w.writerow([])
        w.writerow(["SO SÁNH CA NGÀY / CA ĐÊM"])
        w.writerow(["Ca", "Số lượt bất thường", "Tổng thời gian (phút)",
                    "Số xe-ngày", "Tỷ lệ bất thường (%)"])
        for agg in daynight_split(day_results, denom_hours):
            w.writerow([agg.label, agg.abnormal_count, agg.abnormal_min,
                        agg.car_days, agg.abnormal_rate])

        # Mẫu theo thứ
        w.writerow([])
        w.writerow(["MẪU THEO THỨ TRONG TUẦN"])
        w.writerow(["Thứ", "Số ngày", "Số lượt bất thường", "Số xe-ngày", "Tỷ lệ bất thường (%)"])
        for ws_stat in weekday_pattern(day_results, denom_hours):
            w.writerow([ws_stat.name, ws_stat.num_days, ws_stat.abnormal_count,
                        ws_stat.car_days, ws_stat.abnormal_rate])

        # Phân nhóm mức độ nặng
        w.writerow([])
        w.writerow(["PHÂN NHÓM MỨC ĐỘ NẶNG (theo thời gian dừng)"])
        w.writerow(["Nhóm", "Số lượt", "Tổng thời gian (phút)"])
        for b in severity_buckets(day_results):
            w.writerow([b.label, b.count, b.abnormal_min])

        # Điều phối API
        w.writerow([])
        w.writerow(["ĐIỀU PHỐI API (Tasks API logs)"])
        w.writerow([
            "Ngày", "taskCreate", "Poll", "Task unique", "Gán xe", "Lỗi API", "Điểm hot",
        ])
        for d in sorted(day_results, key=lambda x: x.base_date):
            api = d.api_log_stats if isinstance(d.api_log_stats, DayApiLogStats) else None
            if api is None or not api.has_data:
                continue
            w.writerow([
                d.base_date.isoformat(), api.create_count, api.poll_count,
                api.unique_tasks, api.assigned_car_count, api.api_error_count,
                api.hot_point,
            ])

        w.writerow([])
        w.writerow(["ĐỐI CHIẾU CSV ↔ API"])
        w.writerow([
            "Ngày", "CSV task_count", "API create_count", "Chênh (CSV−API)", "Có CSV", "Có API",
        ])
        for r in csv_api_crosscheck(day_results):
            w.writerow([
                r["date"].isoformat(),  # type: ignore[union-attr]
                r["csv_task_count"], r["api_create_count"], r["diff_csv_minus_api"],
                "Có" if r["has_csv"] else "Không",
                "Có" if r["has_api"] else "Không",
            ])

    return out_path
